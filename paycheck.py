import pdf
from enum import auto, IntEnum
from PyPDF2 import PdfFileReader
import openpyxl 

class Paycheck(pdf.Pdf):

	paycheck_enum_extended = False

	def isType(file_path):
		pdf_reader = PdfFileReader(open(file_path, "rb"))

		for x in range(0, pdf_reader.getNumPages()):
			page = pdf_reader.getPage(x)
			buffer_page_text = page.extractText()[:10]
			if(buffer_page_text.find("Paycheck P") != -1):
				return True
		return False

	def __init__(self, input_filepath = "<EMPTY>"):
		super().__init__(input_filepath)
		if(Paycheck.paycheck_enum_extended == False):
			extra_fields = [m.name for m in Paycheck.Fields] + [
				'OMB_NUMBER',
				'EXP_DATE',
				'T_NAME',
				'B_NAME',
				'B_ADDRESS',
				'B_TIN',
				'B_PHONE',
				'P_CONTACT',
				'EMAIL',
				'AVG_PAYROLL',
				'EIDL_LOANREQ',
				'EMP_COUNT',
				'OWNER1_NAME',
				'OWNER1_TITLE',
				'OWNER1_OWNERSHIP',
				'OWNER1_TIN',
				'OWNER1_ADDRESS',
				'OWNER2_NAME',
				'OWNER2_TITLE',
				'OWNER2_OWNERSHIP',
				'OWNER2_TIN',
				'OWNER2_ADDRESS'
			]
			Paycheck.Fields = IntEnum('Paycheck.Fields', extra_fields, start = 0)
			Paycheck.paycheck_enum_extended = True
		
		for x in range((Paycheck.Fields.FILENAME.value), (Paycheck.Fields.OWNER2_ADDRESS.value)):
			self.list_fields.append("<EMPTY>")

	def displayEnum(self):
		for x in Paycheck.Fields:
			print(x.value," ",x.name)
		print("\n")

	def display(self):
		for x in range(0, len(self.list_fields)):
			print("Paycheck-FIELD-[",Paycheck.Fields(x), "] -> ", self.list_fields[x])
		print("\n")

	def extractData(self):
		pdf_reader = PdfFileReader(open(self.list_fields[Paycheck.Fields.FILENAME.value], "rb"))
		
		dictionary = pdf_reader.getFormTextFields()
		list_of_dict_values = []
		for value in dictionary.values(): 
			list_of_dict_values.append(value)

		buffer_page_text = ""
		
		page = pdf_reader.getPage(0)
		buffer_page_text = page.extractText()
		
		#extracts Omb number
		self.list_fields[Paycheck.Fields.OMB_NUMBER.value] = buffer_page_text[-41:-30]
	
		#extracts Exp date
		self.list_fields[Paycheck.Fields.EXP_DATE.value] = buffer_page_text[-12:-1]
		self.list_fields[Paycheck.Fields.EXP_DATE.value] = self.list_fields[Paycheck.Fields.EXP_DATE.value].strip('\n')
		
		#extracts trade name
		self.list_fields[Paycheck.Fields.T_NAME.value] = str(list_of_dict_values[35]) + " " + str(list_of_dict_values[1])
	
		#extract business name
		self.list_fields[Paycheck.Fields.B_NAME.value] = str(list_of_dict_values[0])
		
		#extract business address
		self.list_fields[Paycheck.Fields.B_ADDRESS.value] = str(list_of_dict_values[2]) + " " + str(list_of_dict_values[5])
		
		#extract business TIN
		self.list_fields[Paycheck.Fields.B_TIN.value] = str(list_of_dict_values[3])

		#extracts business phone
		self.list_fields[Paycheck.Fields.B_PHONE.value] = str(list_of_dict_values[4])

		#extracts all other fields
		i = 0
		for x in range(6, 11):
			self.list_fields[Paycheck.Fields.P_CONTACT.value + i] = str(list_of_dict_values[x])
			i += 1

		i = 0
		for x in range(25, 35):
			self.list_fields[Paycheck.Fields.OWNER1_NAME.value + i] = str(list_of_dict_values[x])
			i += 1

		#replaces 'None' in list
		self.list_fields = [sub.replace("None", "") for sub in self.list_fields]
	
def writeToExcel(self, path):
		wb = openpyxl.load_workbook(path)
		sheet = wb.active
		
		r = sheet.max_row + 1
		for c in range(2, len(self.list_fields) + 2):
			c1 = sheet.cell(row = r, column = c)
			c1.value = self.list_fields[c - 2]

		wb.save(path)
		print("Written successfully to -> ",path)

