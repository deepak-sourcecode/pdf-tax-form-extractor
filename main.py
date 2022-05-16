import os
import openpyxl 

import pdf
import formTaxReturn
import f940
import f941
import f944
import paycheck

def excelInit():
	
	r_path = "Data/FormColumnNames.xlsx"
	r_wb = openpyxl.load_workbook(r_path)
	r_sheet = r_wb.active 	

	w_wb = openpyxl.Workbook()
	w_sheet = w_wb.active

	print("MAX ROWS IS -> ", r_sheet.max_row)
	print("MAX COL IS -> ", r_sheet.max_column)

	temp_filename = ""
	for r in range(1, r_sheet.max_row + 1):
		r_cell = r_sheet.cell(row = r, column = 1)
		temp_filename = r_cell.value
		for c in range(2, r_sheet.max_column + 1):
			r_cell = r_sheet.cell(row = r, column = c)
			w_cell = w_sheet.cell(row = 1, column = c - 1)
			w_cell.value = r_cell.value
		w_wb.save("OutputData/"+str(temp_filename)+".xlsx")
		
#writes sr.no to all sheets
def excelShutDown():
	pass

# Main
if __name__ == "__main__":
	os.system("tput reset")
	excelInit()

	list_files  = os.listdir('InputFiles/')
	files_count = len(list_files)

	for x in range(0, files_count):
		pdf_file_path = 'InputFiles/'+list_files[x]
		excel_file_path = ""

		if f940.F940.isType(pdf_file_path):
			obj = f940.F940(pdf_file_path)
			excel_file_path = "OutputData/940.xlsx"
		elif f941.F941.isType(pdf_file_path):
			obj = f941.F941(pdf_file_path)
			excel_file_path = "OutputData/941.xlsx"
		elif f944.F944.isType(pdf_file_path):
			obj = f944.F944(pdf_file_path)
			excel_file_path = "OutputData/944.xlsx"
		elif paycheck.Paycheck.isType(pdf_file_path):
			obj = paycheck.Paycheck(pdf_file_path)
			excel_file_path = "OutputData/Paycheck.xlsx"
		else:
			continue

		print("Extracting->",pdf_file_path)
		obj.extractData()
		obj.writeToExcel(excel_file_path)