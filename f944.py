import formTaxReturn
from enum import auto, IntEnum
from PyPDF2 import PdfFileReader
import openpyxl 

class F944(formTaxReturn.FormTaxReturn):

	f944_enum_extended = False

	def isType(file_path):
		pdf_reader = PdfFileReader(open(file_path, "rb"))

		for x in range(0, pdf_reader.getNumPages()):
			page = pdf_reader.getPage(x)
			buffer_page_text = page.extractText()[:10]
			if(buffer_page_text.find("Form 944") != -1):
				return True
		return False

	def __init__(self, input_filepath = "<EMPTY>"):
		super().__init__(input_filepath)
		if(F944.f944_enum_extended == False):
			extra_fields = [m.name for m in F944.Fields] + [
				'PART1_1',
				'PART1_2',
				'PART1_4a',
				'PART1_4b',
				'PART1_4c',
				'PART1_4d',
				'PART1_4e',
				'PART1_5',
				'PART1_6',
				'PART1_7',
				'PART1_8',
				'PART1_9',
				'PART1_10',
				'PART1_11',
				'PART1_12',
				'PART2_13a',
				'PART2_13b',
				'PART2_13c',
				'PART2_13d',
				'PART2_13e',
				'PART2_13f',
				'PART2_13g',
				'PART2_13h',
				'PART2_13i',
				'PART2_13j',
				'PART2_13k',
				'PART2_13l',
				'PART2_13m',
				'PART3_14',
				'PART4_DNAME',
				'PART4_DPHONE',
				'PART4_PIN',
				'PART5_NAME',
				'PART5_TITLE',
				'PART5_DAYPHONE',
				'PRE_PNAME',
				'PRE_PTIN',
				'PRE_FNAME',
				'PRE_EIN',
				'PRE_ADDRESS',
				'PRE_PHONE',
				'PRE_CITY',
				'PRE_STATE',
				'PRE_ZIP',
				'V_EIN',
				'V_DOLLARS',
				'V_CENTS',
				'V_BNAME',
				'V_ADDRESS',
				'V_FADDRESS'	
			]
			F944.Fields = IntEnum('F944.Fields', extra_fields, start = 0)
			F944.f944_enum_extended = True
		
		for x in range((F944.Fields.FADDRESS.value), (F944.Fields.V_FADDRESS.value)):
			self.list_fields.append("<EMPTY>")

	def displayEnum(self):
		for x in F944.Fields:
			print(x.value," ",x.name)
		print("\n")

	def display(self):
		for x in range(0, len(self.list_fields)):
			print("F944-FIELD-[",F944.Fields(x), "] -> ", self.list_fields[x])
		print("\n")

	def extractData(self):
		super().extractData()
		pdf_reader = PdfFileReader(open(self.list_fields[F944.Fields.FILENAME.value], "rb"))
		
		dictionary = pdf_reader.getFormTextFields()
		list_of_dict_values = []
		for value in dictionary.values(): 
			list_of_dict_values.append(value)

		buffer_page_text = ""
		
		#extracts YEAR
		page = pdf_reader.getPage(0)
		buffer_page_text = page.extractText()
		self.list_fields[F944.Fields.YEAR.value] = buffer_page_text[:17]

		#extracts part1
		self.list_fields[F944.Fields.PART1_1.value] = str(list_of_dict_values[18])+ "." +str(list_of_dict_values[19])
		self.list_fields[F944.Fields.PART1_2.value] = str(list_of_dict_values[20])+ "." +str(list_of_dict_values[21])		 
		self.list_fields[F944.Fields.PART1_4a.value] = str(list_of_dict_values[24])+ "." +str(list_of_dict_values[25])		 
		self.list_fields[F944.Fields.PART1_4b.value] = str(list_of_dict_values[28])+ "." +str(list_of_dict_values[29])		 
		self.list_fields[F944.Fields.PART1_4c.value] = str(list_of_dict_values[32])+ "." +str(list_of_dict_values[33])		 
		self.list_fields[F944.Fields.PART1_4d.value] = str(list_of_dict_values[36])+ "." +str(list_of_dict_values[37])

		i = 0
		for x in range(38, 55, 2):
			self.list_fields[F944.Fields.PART1_4e.value + i] = str(list_of_dict_values[x]) + "." + str(list_of_dict_values[x+1])
			i += 1

		#extracts part2
		i = 0
		for x in range(58, 83, 2):
			self.list_fields[F944.Fields.PART2_13a.value + i] = str(list_of_dict_values[x]) + "." + str(list_of_dict_values[x+1])
			i += 1

		
		#extracts part3
		self.list_fields[F944.Fields.PART3_14.value] = str(list_of_dict_values[84])
		
		#extracts part4
		self.list_fields[F944.Fields.PART4_DNAME.value] = str(list_of_dict_values[85])
		self.list_fields[F944.Fields.PART4_DPHONE.value] = str(list_of_dict_values[86])
		self.list_fields[F944.Fields.PART4_PIN.value] = str(list_of_dict_values[87])+str(list_of_dict_values[88])+str(list_of_dict_values[89])+str(list_of_dict_values[90])+str(list_of_dict_values[91])


		#extracts rest of the fields
		i = 0
		for x in range(92, len(list_of_dict_values)):
			self.list_fields[F944.Fields.PART5_NAME.value + i] = str(list_of_dict_values[x])
			i += 1

		#replaces 'None' in list
		self.list_fields = [sub.replace("None", "") for sub in self.list_fields]
	
		
	def writeToExcel(self, path):
		wb = openpyxl.load_workbook(path)
		sheet = wb.active
		
		r = sheet.max_row + 1
		for c in range(2, len(self.list_fields) + 2):
			c1 = sheet.cell(row = r, column = c)
			c1.value = self.list_fields[c - 2]

		wb.save(path)
		print("Written successfully to -> ",path)

