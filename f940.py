import formTaxReturn
from enum import auto, IntEnum
from PyPDF2 import PdfFileReader
import openpyxl 

class F940(formTaxReturn.FormTaxReturn):

	f940_enum_extended = False

	def isType(file_path):
		pdf_reader = PdfFileReader(open(file_path, "rb"))
		for x in range(0, pdf_reader.getNumPages()):
			page = pdf_reader.getPage(x)
			buffer_page_text = page.extractText()[:10]
			if(buffer_page_text.find("Form   940") != -1):
				return True
		return False

	def __init__(self, input_filepath = "<EMPTY>"):
		super().__init__(input_filepath)
		if(F940.f940_enum_extended == False):
			extra_fields = [m.name for m in F940.Fields] + [
				'PART2_3',
				'PART2_4',
				'PART2_5',
				'PART2_6',
				'PART2_7',
				'PART2_8',
				'PART3_9',
				'PART3_10',
				'PART3_11',
				'PART4_12',
				'PART4_13',
				'PART4_14',
				'PART4_15',
				'PART5_16a',
				'PART5_16b',
				'PART5_16c',
				'PART5_16d',
				'PART5_17',
				'PART6_DNAME',
				'PART6_DPHONE',
				'PART6_PIN',
				'PART7_NAME',
				'PART7_TITLE',
				'PART7_DAYPHONE',
				'PRE_PNAME',
				'PRE_PTIN',
				'PRE_FNAME',
				'PRE_EIN',
				'PRE_ADDRESS',
				'PRE_PHONE',
				'PRE_CITY',
				'PRE_STATE',
				'PRE_ZIP',
				'V_EIN',
				'V_DOLLARS',
				'V_CENTS',
				'V_BNAME',
				'V_ADDRESS',
				'V_FADDRESS'	
			]
			F940.Fields = IntEnum('F940.Fields', extra_fields, start = 0)
			F940.f940_enum_extended = True
		
		for x in range((F940.Fields.FADDRESS.value), (F940.Fields.V_FADDRESS.value)):
			self.list_fields.append("<EMPTY>")

	def displayEnum(self):
		for x in F940.Fields:
			print(x.value," ",x.name)
		print("\n")

	def display(self):
		for x in range(0, len(self.list_fields)):
			print("F940-FIELD-[",F940.Fields(x), "] -> ", self.list_fields[x])
		print("\n")

	def extractData(self):
		super().extractData()
		pdf_reader = PdfFileReader(open(self.list_fields[F940.Fields.FILENAME.value], "rb"))
		
		dictionary = pdf_reader.getFormTextFields()
		list_of_dict_values = []
		for value in dictionary.values(): 
			list_of_dict_values.append(value)

		buffer_page_text = ""
		
		#extracts YEAR
		page = pdf_reader.getPage(0)
		buffer_page_text = page.extractText()
		self.list_fields[F940.Fields.YEAR.value] = buffer_page_text[:19]

		#extracts part2, part3, part4
		i = 0
		for x in range(20, 45, 2):
			self.list_fields[F940.Fields.PART2_3.value + i] = str(list_of_dict_values[x]) + "." + str(list_of_dict_values[x+1])
			i += 1

		#extracts part5
		i = 0
		for x in range(48, 57, 2):
			self.list_fields[F940.Fields.PART5_16a.value + i] = str(list_of_dict_values[x]) + "." + str(list_of_dict_values[x+1])
			i += 1

		#extracts part6 and rest of the fields
		i = 0
		for x in range(58, len(list_of_dict_values)):
			self.list_fields[F940.Fields.PART6_DNAME.value + i] = str(list_of_dict_values[x])
			i += 1

		#replaces 'None' in list
		self.list_fields = [sub.replace("None", "") for sub in self.list_fields]


	def writeToExcel(self, path):
		wb = openpyxl.load_workbook(path)
		sheet = wb.active
		
		r = sheet.max_row + 1
		for c in range(2, len(self.list_fields) + 2):
			c1 = sheet.cell(row = r, column = c)
			c1.value = self.list_fields[c - 2]

		wb.save(path)
		print("Written successfully to -> ",path)
