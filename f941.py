import formTaxReturn
from enum import auto, IntEnum
from PyPDF2 import PdfFileReader
import openpyxl 

class F941(formTaxReturn.FormTaxReturn):

	f941_enum_extended = False

	def isType(file_path):
		pdf_reader = PdfFileReader(open(file_path, "rb"))

		for x in range(0, pdf_reader.getNumPages()):
			page = pdf_reader.getPage(x)
			buffer_page_text = page.extractText()[:10]
			if(buffer_page_text.find("Form   941") != -1):
				return True
		return False

	def __init__(self, input_filepath = "<EMPTY>"):
		super().__init__(input_filepath)
		if(F941.f941_enum_extended == False):
			extra_fields = [m.name for m in F941.Fields] + [
				'PART1_1',
				'PART1_2',
				'PART1_3',
				'PART1_5a',
				'PART1_5b',
				'PART1_5c',
				'PART1_5d',
				'PART1_5e',
				'PART1_5f',
				'PART1_6',
				'PART1_7',
				'PART1_8',
				'PART1_9',
				'PART1_10',
				'PART1_11',
				'PART1_12',
				'PART1_13',
				'PART1_14',
				'PART1_15',
				'PART2_M1',
				'PART2_M2',
				'PART2_M3',
				'PART2_TOTAL_L',
				'PART3_DATE',
				'PART4_DNAME',
				'PART4_DPHONE',
				'PART4_PIN',
				'PART5_NAME',
				'PART5_TITLE',
				'PART5_DAYPHONE',
				'PRE_PNAME',
				'PRE_PTIN',
				'PRE_FNAME',
				'PRE_EIN',
				'PRE_ADDRESS',
				'PRE_PHONE',
				'PRE_CITY',
				'PRE_STATE',
				'PRE_ZIP',
				'V_EIN',
				'V_DOLLARS',
				'V_CENTS',
				'V_BNAME',
				'V_ADDRESS',
				'V_FADDRESS'	
			]
			F941.Fields = IntEnum('F941.Fields', extra_fields, start = 0)
			F941.f941_enum_extended = True
		
		for x in range((F941.Fields.FADDRESS.value), (F941.Fields.V_FADDRESS.value)):
			self.list_fields.append("<EMPTY>")

	def displayEnum(self):
		for x in F941.Fields:
			print(x.value," ",x.name)
		print("\n")

	def display(self):
		for x in range(0, len(self.list_fields)):
			print("F941-FIELD-[",F941.Fields(x), "] -> ", self.list_fields[x])
		print("\n")

	def extractData(self):
		super().extractData()
		pdf_reader = PdfFileReader(open(self.list_fields[F941.Fields.FILENAME.value], "rb"))
		
		dictionary = pdf_reader.getFormTextFields()
		list_of_dict_values = []
		for value in dictionary.values(): 
			list_of_dict_values.append(value)

		buffer_page_text = ""
		
		#extracts YEAR
		page = pdf_reader.getPage(2)
		buffer_page_text = page.extractText()
		self.list_fields[F941.Fields.YEAR.value] = buffer_page_text[:19]

		#extracts part1
		self.list_fields[F941.Fields.PART1_1.value] = str(list_of_dict_values[18])
		self.list_fields[F941.Fields.PART1_2.value] = str(list_of_dict_values[19])+ "." +str(list_of_dict_values[20])
		self.list_fields[F941.Fields.PART1_3.value] = str(list_of_dict_values[21])+ "." +str(list_of_dict_values[22])
		self.list_fields[F941.Fields.PART1_5a.value] = str(list_of_dict_values[25])+ "." +str(list_of_dict_values[26])
		self.list_fields[F941.Fields.PART1_5b.value] = str(list_of_dict_values[29])+ "." +str(list_of_dict_values[30])
		self.list_fields[F941.Fields.PART1_5c.value] = str(list_of_dict_values[33])+ "." +str(list_of_dict_values[34])
		self.list_fields[F941.Fields.PART1_5d.value] = str(list_of_dict_values[37])+ "." +str(list_of_dict_values[38])
		
		i = 0
		for x in range(39, 62, 2):
			self.list_fields[F941.Fields.PART1_5e.value + i] = str(list_of_dict_values[x]) + "." + str(list_of_dict_values[x+1])
			i += 1

		#extracts part2
		i = 0
		for x in range(65, 72, 2):
			self.list_fields[F941.Fields.PART2_M1.value + i] = str(list_of_dict_values[x]) + "." + str(list_of_dict_values[x+1])
			i += 1

		#extracts part3
		self.list_fields[F941.Fields.PART3_DATE.value] = str(list_of_dict_values[73])

		#extracts part4
		self.list_fields[F941.Fields.PART4_DNAME.value] = str(list_of_dict_values[74])
		self.list_fields[F941.Fields.PART4_DPHONE.value] = str(list_of_dict_values[75])
		self.list_fields[F941.Fields.PART4_PIN.value] = str(list_of_dict_values[76])+str(list_of_dict_values[77])+str(list_of_dict_values[78])+str(list_of_dict_values[79])+str(list_of_dict_values[80])

		#extracts rest of the fields
		i = 0
		for x in range(81, len(list_of_dict_values)):
			self.list_fields[F941.Fields.PART5_NAME.value + i] = str(list_of_dict_values[x])
			i += 1

		#replaces 'None' in list
		self.list_fields = [sub.replace("None", "") for sub in self.list_fields]
	

	def writeToExcel(self, path):
		wb = openpyxl.load_workbook(path)
		sheet = wb.active
		
		r = sheet.max_row + 1
		for c in range(2, len(self.list_fields) + 2):
			c1 = sheet.cell(row = r, column = c)
			c1.value = self.list_fields[c - 2]

		wb.save(path)
		print("Written successfully to -> ",path)
